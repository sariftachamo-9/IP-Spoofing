import flet as ft
import openpyxl
import os
import datetime

# ---------------- CONFIG & PATHS ----------------
base_dir = os.path.join(os.path.expanduser("~"), "Documents", "TT_Academy")
os.makedirs(base_dir, exist_ok=True)

STUDENT_FILE = os.path.join(base_dir, "students.xlsx")
ORDER_FILE = os.path.join(base_dir, "orders.xlsx")
MENU_FILE = os.path.join(base_dir, "menu.xlsx")
INVENTORY_BAR_FILE = os.path.join(base_dir, "inventory_bar.xlsx")
INVENTORY_COFFEE_FILE = os.path.join(base_dir, "inventory_coffee.xlsx")
RECEIPT_DIR = os.path.join(base_dir, "receipts")
os.makedirs(RECEIPT_DIR, exist_ok=True)

# ---------------- DATA MANAGER ----------------
class DataManager:
    def __init__(self):
        self.menu = {}
        self.cart = []
        self.init_files()
        self.load_menu_data()

    def init_files(self):
        # Sheet initialization logic
        files_to_init = {
            STUDENT_FILE: ["Name","Contact","Location","Total","Paid","Pending","Payment Mode","Remarks"],
            ORDER_FILE: ["BillID","Date","Item","Qty","Price","Total"],
            MENU_FILE: ["Item","Price"],
            INVENTORY_BAR_FILE: ["Item","Unit (ml)","Current Stock","Used","Restocked","Last Updated"],
            INVENTORY_COFFEE_FILE: ["Item","Unit","Unit Type","Current Stock","Used","Restocked","Last Updated"]
        }
        for path, header in files_to_init.items():
            if not os.path.exists(path):
                wb = openpyxl.Workbook(); ws = wb.active; ws.title="Sheet1"
                # For specific files, set correct sheet titles as expected by legacy app
                if path == MENU_FILE: ws.title = "Menu"
                elif "inventory" in path: ws.title = "Inventory"
                elif path == STUDENT_FILE: ws.title = "Students"
                elif path == ORDER_FILE: ws.title = "Orders"
                ws.append(header)
                if path == MENU_FILE:
                    for d in [("Beer",350),("Vodka",450),("Whiskey",600),("Latte",250),("Cappuccino",300)]: ws.append(d)
                if path == INVENTORY_BAR_FILE:
                    for d in [("Beer",650,100,0,0,"-"),("Vodka",750,50,0,0,"-"),("Whiskey",750,30,0,0,"-")]: ws.append(d)
                wb.save(path)

    def load_menu_data(self):
        self.menu.clear()
        wb = openpyxl.load_workbook(MENU_FILE); ws = wb["Menu"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]: self.menu[row[0]] = row[1]
        wb.close()

    def save_item_sold(self, bill_id, date_str, item):
        wb = openpyxl.load_workbook(ORDER_FILE); ws = wb["Orders"]
        ws.append([bill_id, date_str, item['name'], item['qty'], item['price'], item['total']])
        wb.save(ORDER_FILE); wb.close()

    def update_inventory_stock(self, item_name, qty):
        # Deduct from Bar
        wb = openpyxl.load_workbook(INVENTORY_BAR_FILE); ws = wb["Inventory"]
        deducted = False
        for row in ws.iter_rows(min_row=2):
            if row[0].value == item_name:
                row[2].value = (row[2].value or 0) - qty
                row[3].value = (row[3].value or 0) + qty
                row[5].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                deducted = True; break
        wb.save(INVENTORY_BAR_FILE)
        if deducted: return
        
        # Deduct from Coffee
        wb = openpyxl.load_workbook(INVENTORY_COFFEE_FILE); ws = wb["Inventory"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == item_name:
                row[3].value = (row[3].value or 0) - qty
                row[4].value = (row[4].value or 0) + qty
                row[6].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                break
        wb.save(INVENTORY_COFFEE_FILE)

# ---------------- UI COMPONENTS ----------------

class MobileApp:
    def __init__(self, page: ft.Page):
        self.page = page
        self.dm = DataManager()
        self.setup_page()
        self.build_ui()

    def setup_page(self):
        self.page.title = "Barista POS Mobile"
        self.page.theme_mode = ft.ThemeMode.DARK
        self.page.padding = 0
        self.page.window_width = 400
        self.page.window_height = 800

    def build_ui(self):
        self.cart_list_view = ft.ListView(expand=True, spacing=10, padding=15)
        self.pos_view = self.create_pos_view()
        self.students_view = self.create_students_view()
        self.inventory_view = self.create_inventory_view()
        self.menu_view = self.create_menu_view()
        self.history_view = self.create_history_view()

        self.views = {
            0: self.pos_view,
            1: self.students_view,
            2: self.inventory_view,
            3: self.menu_view,
            4: self.history_view
        }

        self.container = ft.Container(content=self.pos_view, expand=True, padding=0)

        self.refresh_students_list()
        self.refresh_all_data()

        self.page.navigation_bar = ft.NavigationBar(
            destinations=[
                ft.NavigationBarDestination(icon=ft.icons.SHOPPING_CART_OUTLINED, selected_icon=ft.icons.SHOPPING_CART, label="POS"),
                ft.NavigationBarDestination(icon=ft.icons.PERSON_OUTLINED, selected_icon=ft.icons.PERSON, label="Students"),
                ft.NavigationBarDestination(icon=ft.icons.INVENTORY_2_OUTLINED, selected_icon=ft.icons.INVENTORY_2, label="Stock"),
                ft.NavigationBarDestination(icon=ft.icons.RESTAURANT_MENU, selected_icon=ft.icons.RESTAURANT_MENU, label="Menu"),
                ft.NavigationBarDestination(icon=ft.icons.HISTORY, label="History"),
            ],
            on_change=self.handle_nav_change,
            selected_index=0
        )

        self.page.add(
            ft.AppBar(
                title=ft.Text("T&T ACADEMY", weight="bold", size=20),
                center_title=True,
                bgcolor=ft.colors.SURFACE_VARIANT,
                elevation=2,
            ),
            self.container
        )

    def handle_nav_change(self, e):
        index = e.control.selected_index
        self.container.content = self.views[index]
        # Refresh data when switching tabs for best "live" experience
        if index == 1: self.refresh_students_list()
        elif index == 2: self.refresh_stock()
        elif index == 3: self.refresh_menu()
        elif index == 4: self.refresh_history()
        self.page.update()

    # --- POS VIEW ---
    def create_pos_view(self):
        self.grand_total_text = ft.Text("Rs 0", size=24, weight="bold", color=ft.colors.GREEN_ACCENT)
        
        # Product Category Filter (Optional refinement)
        product_grid = ft.Row(wrap=True, spacing=10, scroll=ft.ScrollMode.HIDDEN)
        
        def add_item_from_card(name, price):
            # Default qty 1 when clicking a card
            self.dm.cart.append({
                'name': name,
                'qty': 1,
                'price': price,
                'total': price
            })
            self.refresh_cart()
            self.page.snack_bar = ft.SnackBar(ft.Text(f"Added {name}"), duration=1000)
            self.page.snack_bar.open = True
            self.page.update()

        for name, price in self.dm.menu.items():
            product_grid.controls.append(
                ft.GestureDetector(
                    on_tap=lambda _, n=name, p=price: add_item_from_card(n, p),
                    content=ft.Container(
                        content=ft.Column([
                            ft.Icon(ft.icons.FASTFOOD, size=30, color=ft.colors.PRIMARY),
                            ft.Text(name, weight="bold", size=12, text_align="center", max_lines=1),
                            ft.Text(f"Rs {price}", size=11, color=ft.colors.GREY_400),
                        ], horizontal_alignment="center", spacing=5),
                        width=100,
                        height=100,
                        bgcolor=ft.colors.SURFACE_VARIANT,
                        border_radius=15,
                        padding=10,
                        alignment=ft.alignment.center
                    )
                )
            )

        return ft.Column([
            ft.Text("Quick Select Menu", size=18, weight="bold"),
            ft.Container(content=product_grid, height=130),
            ft.Divider(height=10),
            ft.Text("Items in Cart", size=18, weight="w500"),
            self.cart_list_view,
            ft.Container(
                padding=20,
                bgcolor=ft.colors.SURFACE_VARIANT,
                border_radius=ft.border_radius.only(top_left=20, top_right=20),
                content=ft.Column([
                    ft.Row([
                        ft.Text("Grand Total:", size=18),
                        self.grand_total_text
                    ], alignment="spaceBetween"),
                    ft.ElevatedButton(
                        "Process Bill", 
                        icon=ft.icons.CHECK_CIRCLE, 
                        on_click=self.process_bill,
                        style=ft.ButtonStyle(bgcolor=ft.colors.PRIMARY, color=ft.colors.WHITE, shape=ft.RoundedRectangleBorder(radius=15)),
                        height=55,
                        width=float("inf")
                    )
                ], spacing=15)
            )
        ], expand=True)

    def refresh_cart(self):
        self.cart_list_view.controls.clear()
        total = 0
        for i, item in enumerate(self.dm.cart):
            total += item['total']
            
            # Stepper UI
            def change_qty(idx, delta):
                new_qty = self.dm.cart[idx]['qty'] + delta
                if new_qty > 0:
                    self.dm.cart[idx]['qty'] = new_qty
                    self.dm.cart[idx]['total'] = new_qty * self.dm.cart[idx]['price']
                    self.refresh_cart()
                elif new_qty == 0:
                    self.remove_cart_item(idx)

            self.cart_list_view.controls.append(
                ft.Card(
                    elevation=1,
                    content=ft.Container(
                        padding=10,
                        content=ft.Row([
                            ft.Icon(ft.icons.RESTAURANT, color=ft.colors.PRIMARY_CONTAINER),
                            ft.Column([
                                ft.Text(item['name'], weight="bold"),
                                ft.Text(f"Rs {item['price']} each", size=12, color=ft.colors.GREY_400),
                            ], expand=True),
                            ft.Row([
                                ft.IconButton(ft.icons.REMOVE_CIRCLE_OUTLINE, icon_size=20, on_click=lambda _, idx=i: change_qty(idx, -1)),
                                ft.Text(str(item['qty']), weight="bold", size=16),
                                ft.IconButton(ft.icons.ADD_CIRCLE_OUTLINE, icon_size=20, on_click=lambda _, idx=i: change_qty(idx, 1)),
                            ], spacing=0)
                        ])
                    )
                )
            )
        self.grand_total_text.value = f"Rs {total}"
        self.page.update()

    def remove_cart_item(self, idx):
        self.dm.cart.pop(idx)
        self.refresh_cart()

    def process_bill(self, _):
        if not self.dm.cart: return
        now = datetime.datetime.now()
        bill_id = now.strftime("%Y%m%d%H%M%S")
        date_str = now.strftime("%Y-%m-%d %H:%M:%S")
        
        for item in self.dm.cart:
            self.dm.save_item_sold(bill_id, date_str, item)
            self.dm.update_inventory_stock(item['name'], item['qty'])
        
        # Show Success
        self.page.snack_bar = ft.SnackBar(ft.Text(f"Bill Generated! ID: {bill_id}"))
        self.page.snack_bar.open = True
        self.dm.cart.clear()
        self.refresh_cart()
        self.refresh_history()
        self.refresh_stock()

    # --- STUDENTS VIEW ---
    def create_students_view(self):
        self.stud_search = ft.TextField(
            label="Search Students", 
            prefix_icon=ft.icons.SEARCH, 
            on_change=self.filter_students,
            border_radius=15,
            filled=True,
            hint_text="Name or Number..."
        )
        self.stud_list = ft.ListView(expand=True, spacing=10, padding=10)
        return ft.Column([
            ft.Padding(padding=ft.padding.only(left=10, right=10, top=10), content=self.stud_search),
            self.stud_list,
            ft.Padding(padding=10, content=ft.ElevatedButton(
                "Add New Student", 
                icon=ft.icons.PERSON_ADD, 
                on_click=self.show_add_student_dialog,
                width=float("inf"),
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10))
            ))
        ], expand=True)

    def filter_students(self, e):
        term = e.control.value.lower()
        self.refresh_students_list(term)

    def refresh_students_list(self, term=""):
        self.stud_list.controls.clear()
        wb = openpyxl.load_workbook(STUDENT_FILE); ws = wb["Students"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not term or term in str(row[0]).lower() or term in str(row[1]).lower():
                self.stud_list.controls.append(
                    ft.Card(
                        elevation=2,
                        content=ft.ListTile(
                            leading=ft.CircleAvatar(content=ft.Text(str(row[0])[0])),
                            title=ft.Text(str(row[0]), weight="bold"),
                            subtitle=ft.Text(f"📞 {row[1]} \n💸 Pending: Rs {row[5]}"),
                            is_three_line=True,
                            trailing=ft.Icon(ft.icons.CHEVRON_RIGHT),
                            on_click=lambda _, r=row: self.show_student_details(r)
                        )
                    )
                )
        wb.close()
        self.page.update()

    def show_student_details(self, row):
        pay_input = ft.TextField(label="Add Payment Amount", keyboard_type="number")
        self.page.dialog = ft.AlertDialog(
            title=ft.Text(f"Details: {row[0]}"),
            content=ft.Column([
                ft.Text(f"📍 Location: {row[2]}"),
                ft.Text(f"💰 Total Fees: Rs {row[3]}"),
                ft.Text(f"✅ Paid: Rs {row[4]}"),
                ft.Text(f"⚠️ Pending: Rs {row[5]}", color=ft.colors.RED_ACCENT, weight="bold"),
                ft.Divider(),
                pay_input,
            ], tight=True, spacing=10),
            actions=[
                ft.TextButton("Cancel", on_click=lambda _: self.close_dialog()),
                ft.ElevatedButton("Save Payment", on_click=lambda _: self.update_student_payment(row[0], row[1], pay_input.value))
            ],
            actions_alignment="end"
        )
        self.page.dialog.open = True
        self.page.update()

    def update_student_payment(self, name, contact, amount):
        if not amount: return
        try: add = float(amount)
        except: return
        
        wb = openpyxl.load_workbook(STUDENT_FILE); ws = wb["Students"]
        for r in ws.iter_rows(min_row=2):
            if str(r[0].value) == str(name) and str(r[1].value) == str(contact):
                r[4].value = (r[4].value or 0) + add
                r[5].value = (r[3].value or 0) - r[4].value
                break
        wb.save(STUDENT_FILE); wb.close()
        self.close_dialog()
        self.refresh_students_list()
        self.page.snack_bar = ft.SnackBar(ft.Text("Payment Updated Successfully!"))
        self.page.snack_bar.open = True
        self.page.update()

    def show_add_student_dialog(self, _):
        name = ft.TextField(label="Full Name")
        contact = ft.TextField(label="Contact Number")
        loc = ft.TextField(label="Location")
        total = ft.TextField(label="Total Amount", value="0")
        paid = ft.TextField(label="Initial Paid", value="0")
        
        def save_new(_):
            try:
                t = float(total.value); p = float(paid.value)
                wb = openpyxl.load_workbook(STUDENT_FILE); ws = wb["Students"]
                ws.append([name.value, contact.value, loc.value, t, p, t-p, "Cash", "Mobile Entry"])
                wb.save(STUDENT_FILE); wb.close()
                self.close_dialog()
                self.refresh_students_list()
            except: pass

        self.page.dialog = ft.AlertDialog(
            title=ft.Text("Add New Student"),
            content=ft.Column([name, contact, loc, total, paid], tight=True),
            actions=[
                ft.TextButton("Cancel", on_click=lambda _: self.close_dialog()),
                ft.ElevatedButton("Save", on_click=save_new)
            ]
        )
        self.page.dialog.open = True
        self.page.update()

    def close_dialog(self):
        self.page.dialog.open = False
        self.page.update()

    # --- INVENTORY VIEW ---
    def create_inventory_view(self):
        self.stock_search = ft.TextField(label="Search Stock", prefix_icon=ft.icons.SEARCH, on_change=self.filter_stock)
        self.stock_list = ft.ListView(expand=True, spacing=8, padding=10)
        return ft.Column([
            ft.Padding(padding=ft.padding.only(left=10, right=10, top=10), content=self.stock_search),
            self.stock_list
        ], expand=True)

    def filter_stock(self, e):
        term = e.control.value.lower()
        self.refresh_stock(term)

    def refresh_stock(self, term=""):
        self.stock_list.controls.clear()
        # Bar
        wb = openpyxl.load_workbook(INVENTORY_BAR_FILE); ws = wb["Inventory"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not term or term in str(row[0]).lower():
                self.stock_list.controls.append(self.create_stock_card(row, "BAR"))
        wb.close()
        # Coffee
        wb = openpyxl.load_workbook(INVENTORY_COFFEE_FILE); ws = wb["Inventory"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not term or term in str(row[0]).lower():
                self.stock_list.controls.append(self.create_stock_card(row, "COFFEE"))
        wb.close()
        self.page.update()

    def create_stock_card(self, row, category):
        stock = row[2] if category == "BAR" else row[3]
        unit = row[1]
        color = ft.colors.GREEN_ACCENT if stock > 10 else ft.colors.ORANGE_ACCENT if stock > 0 else ft.colors.RED_ACCENT
        
        return ft.Card(
            elevation=2,
            content=ft.Container(
                padding=12,
                content=ft.Row([
                    ft.Icon(ft.icons.LIQUOR if category == "BAR" else ft.icons.COFFEE_MAKER, color=color, size=30),
                    ft.Column([
                        ft.Text(str(row[0]), weight="bold", size=16),
                        ft.Text(f"{category} | Unit: {unit}", size=12, color=ft.colors.GREY_400)
                    ], expand=True),
                    ft.Column([
                        ft.Text(f"{stock}", size=20, weight="bold", color=color),
                        ft.Text("IN STOCK" if stock > 0 else "OUT OF STOCK", size=9, weight="bold", color=color)
                    ], horizontal_alignment="center", spacing=0)
                ])
            )
        )

    # --- MENU VIEW ---
    def create_menu_view(self):
        self.menu_list = ft.ListView(expand=True, spacing=8, padding=10)
        return ft.Column([
            ft.Padding(padding=10, content=ft.Text("Menu Items", size=20, weight="bold")),
            self.menu_list,
            ft.Padding(padding=10, content=ft.ElevatedButton("Add Menu Item", on_click=self.show_add_menu_item))
        ], expand=True)

    def refresh_menu(self):
        self.menu_list.controls.clear()
        self.dm.load_menu_data()
        for k, v in self.dm.menu.items():
            self.menu_list.controls.append(
                ft.Card(
                    content=ft.ListTile(
                        title=ft.Text(k, weight="bold"),
                        subtitle=ft.Text(f"Price: Rs {v}"),
                        trailing=ft.IconButton(ft.icons.DELETE_OUTLINE, on_click=lambda _, item=k: self.delete_menu_item(item))
                    )
                )
            )
        self.item_select.options = [ft.dropdown.Option(k) for k in self.dm.menu.keys()]
        self.page.update()

    def show_add_menu_item(self, _):
        name = ft.TextField(label="Item Name")
        price = ft.TextField(label="Price", keyboard_type="number")
        def save(_):
            self.dm.menu[name.value] = float(price.value)
            # Write to Excel
            wb = openpyxl.load_workbook(MENU_FILE); ws = wb["Menu"]
            ws.append([name.value, float(price.value)])
            wb.save(MENU_FILE); wb.close()
            self.close_dialog()
            self.refresh_menu()
        
        self.page.dialog = ft.AlertDialog(title=ft.Text("Add Menu Item"), content=ft.Column([name, price], tight=True), actions=[ft.ElevatedButton("Save", on_click=save)])
        self.page.dialog.open = True
        self.page.update()

    def delete_menu_item(self, item):
        if item in self.dm.menu:
            del self.dm.menu[item]
            wb = openpyxl.load_workbook(MENU_FILE); ws = wb["Menu"]
            # Find and delete row
            for r in ws.iter_rows(min_row=2):
                if r[0].value == item:
                    ws.delete_rows(r[0].row)
                    break
            wb.save(MENU_FILE); wb.close()
            self.refresh_menu()

    # --- HISTORY VIEW ---
    def create_history_view(self):
        self.history_list = ft.ListView(expand=True, spacing=5, padding=10)
        return ft.Column([
            ft.Padding(padding=10, content=ft.Row([
                ft.Icon(ft.icons.HISTORY, color=ft.colors.BLUE_200),
                ft.Text("Recent Orders", size=20, weight="bold")
            ])),
            self.history_list
        ], expand=True)

    def refresh_history(self):
        self.history_list.controls.clear()
        if not os.path.exists(ORDER_FILE): return
        wb = openpyxl.load_workbook(ORDER_FILE); ws = wb["Orders"]
        # Show last 30
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in reversed(rows[-30:]):
            if not row[0]: continue
            self.history_list.controls.append(
                ft.Card(
                    content=ft.ListTile(
                        leading=ft.Icon(ft.icons.RECEIPT_LONG, color=ft.colors.BLUE_ACCENT),
                        title=ft.Text(f"Bill: {row[0]}"),
                        subtitle=ft.Text(f"{row[2]} x{row[3]} | Rs {row[5]}"),
                        trailing=ft.Text(str(row[1]).split()[0], size=11, color=ft.colors.GREY_400),
                        is_three_line=True,
                        on_click=lambda _, r=row: self.show_order_details(r)
                    )
                )
            )
        wb.close()
        self.page.update()

    def show_order_details(self, row):
        self.page.dialog = ft.AlertDialog(
            title=ft.Text(f"Order #{row[0]}"),
            content=ft.Column([
                ft.Text(f"📅 Date: {row[1]}"),
                ft.Text(f"🍔 Item: {row[2]}"),
                ft.Text(f"🔢 Quantity: {row[3]}"),
                ft.Text(f"💵 Price: Rs {row[4]}"),
                ft.Text(f"💰 Total: Rs {row[5]}", size=18, weight="bold", color=ft.colors.GREEN_ACCENT),
            ], tight=True, spacing=10),
            actions=[ft.TextButton("Close", on_click=lambda _: self.close_dialog())]
        )
        self.page.dialog.open = True
        self.page.update()

    def refresh_all_data(self):
        self.refresh_students_list()
        self.refresh_stock()
        self.refresh_menu()
        self.refresh_history()

def main(page: ft.Page):
    MobileApp(page)

if __name__ == "__main__":
    ft.app(target=main)
