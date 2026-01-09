import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import datetime
import os
import openpyxl

# ---------------- BASE PATHS ----------------
base_dir = os.path.join(os.path.expanduser("~"), "Documents", "TT_Academy")
os.makedirs(base_dir, exist_ok=True)

STUDENT_FILE = os.path.join(base_dir, "students.xlsx")
ORDER_FILE = os.path.join(base_dir, "orders.xlsx")
MENU_FILE = os.path.join(base_dir, "menu.xlsx")
INVENTORY_BAR_FILE = os.path.join(base_dir, "inventory_bar.xlsx")
INVENTORY_COFFEE_FILE = os.path.join(base_dir, "inventory_coffee.xlsx")
RECEIPT_DIR = os.path.join(base_dir, "receipts")
os.makedirs(RECEIPT_DIR, exist_ok=True)

# ---------------- GLOBAL DATA ----------------
MENU = {}
cart = []
last_receipt_file = None
selected_student_row_idx = None
item_var = None

# ---------------- INIT FILES ----------------
def init_files():
    if not os.path.exists(STUDENT_FILE):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Students"
        ws.append(["Name","Contact","Location","Total","Paid","Pending","Payment Mode","Remarks"])
        wb.save(STUDENT_FILE)
    if not os.path.exists(ORDER_FILE):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Orders"
        ws.append(["BillID","Date","Item","Qty","Price","Total"])
        wb.save(ORDER_FILE)
    if not os.path.exists(MENU_FILE):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Menu"
        ws.append(["Item","Price"])
        defaults = [("Beer",350),("Vodka",450),("Whiskey",600),("Latte",250),("Cappuccino",300)]
        for d in defaults: ws.append(d)
        wb.save(MENU_FILE)
    
    # Check Bar Inventory Schema
    if os.path.exists(INVENTORY_BAR_FILE):
        try:
            wb = openpyxl.load_workbook(INVENTORY_BAR_FILE); ws = wb.active
            if ws.cell(row=1, column=2).value != "Unit (ml)":
                wb.close(); os.remove(INVENTORY_BAR_FILE)
            else: wb.close()
        except: pass

    # Check Coffee Inventory Schema
    if os.path.exists(INVENTORY_COFFEE_FILE):
        try:
            wb = openpyxl.load_workbook(INVENTORY_COFFEE_FILE); ws = wb.active
            # New schema: Item, Unit, Unit Type, Stock...
            if ws.cell(row=1, column=2).value != "Unit": # Check Col 2
                wb.close(); os.remove(INVENTORY_COFFEE_FILE)
            else: wb.close()
        except: pass

    # Init Bar Inventory
    if not os.path.exists(INVENTORY_BAR_FILE):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Inventory"
        ws.append(["Item","Unit (ml)","Current Stock","Used","Restocked","Last Updated"])
        for d in [("Beer",650,100,0,0,"-"),("Vodka",750,50,0,0,"-"),("Whiskey",750,30,0,0,"-")]:
            ws.append(d)
        wb.save(INVENTORY_BAR_FILE)

    # Init Coffee Inventory
    if not os.path.exists(INVENTORY_COFFEE_FILE):
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Inventory"
        ws.append(["Item","Unit","Unit Type","Current Stock","Used","Restocked","Last Updated"])
        # Items match MENU items to ensure auto-deduction works
        for d in [("Latte",250,"ml",5000,0,0,"-"),("Cappuccino",250,"ml",5000,0,0,"-")]:
            ws.append(d)
        wb.save(INVENTORY_COFFEE_FILE)

# ---------------- MENU ----------------
def load_menu():
    global MENU, item_var
    MENU.clear()
    wb=openpyxl.load_workbook(MENU_FILE); ws=wb["Menu"]
    for row in ws.iter_rows(min_row=2,values_only=True):
        if row[0]: MENU[row[0]]=row[1]
    refresh_item_menu()
    load_menu_table()
    wb.close()

def save_menu_excel():
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Menu"; ws.append(["Item","Price"])
    for k,v in MENU.items(): ws.append([k,v])
    wb.save(MENU_FILE)

def refresh_item_menu():
    item_menu['menu'].delete(0,'end')
    for item in MENU.keys(): item_menu['menu'].add_command(label=item, command=tk._setit(item_var,item))

def load_menu_table():
    for row in menu_table.get_children(): menu_table.delete(row)
    for k,v in MENU.items(): menu_table.insert("",tk.END,values=(k,v))

def add_item_menu(): 
    n,p = menu_name_entry.get().strip(), menu_price_entry.get().strip()
    if not n or not p: messagebox.showwarning("Error","Name and price required"); return
    try: p=float(p)
    except: messagebox.showerror("Error","Price must be number"); return
    MENU[n]=p; save_menu_excel(); load_menu(); menu_name_entry.delete(0,tk.END); menu_price_entry.delete(0,tk.END)

def edit_item_menu():
    sel=menu_table.selection(); 
    if not sel: messagebox.showwarning("Error","Select item to edit"); return
    old=menu_table.item(sel[0])['values'][0]; new_name, new_price=menu_name_entry.get().strip(), menu_price_entry.get().strip()
    if not new_name or not new_price: messagebox.showwarning("Error","New name and price required"); return
    try: new_price=float(new_price)
    except: messagebox.showerror("Error","Price must be number"); return
    MENU.pop(old); MENU[new_name]=new_price; save_menu_excel(); load_menu(); menu_name_entry.delete(0,tk.END); menu_price_entry.delete(0,tk.END)

def remove_item_menu():
    sel=menu_table.selection(); 
    if not sel: messagebox.showwarning("Error","Select item to remove"); return
    old=menu_table.item(sel[0])['values'][0]; 
    if messagebox.askyesno("Confirm",f"Remove {old}?"): MENU.pop(old); save_menu_excel(); load_menu()

# ---------------- GENERIC INVENTORY FUNCTIONS ----------------
def load_inventory_table(tree, filepath):
    for r in tree.get_children(): tree.delete(r)
    if not os.path.exists(filepath): return
    wb=openpyxl.load_workbook(filepath); ws=wb["Inventory"]
    for row in ws.iter_rows(min_row=2,values_only=True): tree.insert("",tk.END,values=row)
    wb.close()

def add_inventory_item(tree, filepath, name_entry, stock_entry, unit_entry=None, unit_type_var=None):
    name = name_entry.get().strip()
    stock = stock_entry.get().strip()
    
    # Unit Logic
    unit_val = unit_entry.get().strip() if unit_entry else None
    unit_type = unit_type_var.get() if unit_type_var else None

    if not name or not stock: messagebox.showwarning("Error", "Name and Stock required"); return
    
    # Validation per file type
    is_bar = "bar" in filepath.lower()
    is_coffee = "coffee" in filepath.lower()
    
    if is_bar and (not unit_val): messagebox.showwarning("Error", "Unit (ml) required"); return
    if is_coffee and (not unit_val or not unit_type): messagebox.showwarning("Error", "Unit Size & Type required"); return

    try: 
        stock = float(stock)
        if unit_val: unit_val = float(unit_val)
    except: messagebox.showerror("Error", "Stock/Unit must be number"); return
    
    wb = openpyxl.load_workbook(filepath); ws = wb["Inventory"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == name:
            messagebox.showwarning("Error", "Item already exists"); wb.close(); return
    
    if is_bar:
        ws.append([name, unit_val, stock, 0, 0, datetime.datetime.now().strftime("%Y-%m-%d")])
    elif is_coffee:
        ws.append([name, unit_val, unit_type, stock, 0, 0, datetime.datetime.now().strftime("%Y-%m-%d")])
    else:
        # Fallback
        ws.append([name, stock, 0, 0, datetime.datetime.now().strftime("%Y-%m-%d")])
        
    wb.save(filepath); wb.close()
    load_inventory_table(tree, filepath)
    name_entry.delete(0, tk.END); stock_entry.delete(0, tk.END)
    if unit_entry: unit_entry.delete(0, tk.END)
    # Reset drop down? Keep current selection

def edit_inventory_item(tree, filepath, name_entry, stock_entry, unit_entry=None, unit_type_var=None):
    sel = tree.selection()
    if not sel: messagebox.showwarning("Error", "Select item to edit"); return
    old_item_name = tree.item(sel[0])['values'][0]
    
    new_name = name_entry.get().strip()
    new_stock = stock_entry.get().strip()
    new_unit = unit_entry.get().strip() if unit_entry else None
    new_type = unit_type_var.get() if unit_type_var else None
    
    if not new_name and not new_stock and not new_unit and not new_type: 
         messagebox.showwarning("Error", "Enter data to update"); return
    
    is_bar = "bar" in filepath.lower()
    is_coffee = "coffee" in filepath.lower()

    wb = openpyxl.load_workbook(filepath); ws = wb["Inventory"]
    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == old_item_name:
            if new_name: row[0].value = new_name
            
            if is_bar:
                if new_unit: 
                     try: row[1].value = float(new_unit)
                     except: pass
                if new_stock:
                     try: row[2].value = float(new_stock)
                     except: pass
                row[5].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            elif is_coffee:
                if new_unit: 
                     try: row[1].value = float(new_unit)
                     except: pass
                if new_type: row[2].value = new_type
                if new_stock:
                     try: row[3].value = float(new_stock)
                     except: pass
                row[6].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            found = True
            break
    
    if found:
        wb.save(filepath); wb.close()
        load_inventory_table(tree, filepath)
        name_entry.delete(0, tk.END); stock_entry.delete(0, tk.END)
        if unit_entry: unit_entry.delete(0, tk.END)
    else:
        wb.close()
        messagebox.showerror("Error", "Item not found in file")

def delete_inventory_item(tree, filepath):
    sel = tree.selection()
    if not sel: messagebox.showwarning("Error", "Select item to delete"); return
    item_name = tree.item(sel[0])['values'][0]
    if not messagebox.askyesno("Confirm", f"Delete {item_name}?"): return
    wb = openpyxl.load_workbook(filepath); ws = wb["Inventory"]
    idx_to_del = None
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == item_name:
            idx_to_del = idx; break
    if idx_to_del:
        ws.delete_rows(idx_to_del); wb.save(filepath); wb.close(); load_inventory_table(tree, filepath)
    else: wb.close()

def restock_inventory_item(tree, filepath):
    sel=tree.selection()
    if not sel: messagebox.showwarning("Error","Select item to restock"); return
    item_data=tree.item(sel[0])['values']
    item_name=item_data[0]
    
    is_bar = "bar" in filepath.lower()
    is_coffee = "coffee" in filepath.lower()
    
    restock_win = tk.Toplevel(root)
    restock_win.title("Restock Item")
    restock_win.geometry("300x350")
    
    tk.Label(restock_win, text=f"Restock: {item_name}", font=("Helvetica",11,"bold")).pack(pady=10)
    
    unit_var = tk.StringVar()
    type_var = tk.StringVar()

    if is_bar:
        tk.Label(restock_win, text="Unit (ml):").pack()
        tk.Entry(restock_win, textvariable=unit_var).pack(pady=5)
        # Bar Data: Item[0], Unit[1], Stock[2]...
        unit_var.set(str(item_data[1]))
    elif is_coffee:
         tk.Label(restock_win, text="Unit:").pack()
         tk.Entry(restock_win, textvariable=unit_var).pack(pady=5)
         unit_var.set(str(item_data[1]))
         
         tk.Label(restock_win, text="Unit Type:").pack()
         type_combo=ttk.Combobox(restock_win, textvariable=type_var, values=["ml","gm"], state="readonly")
         type_combo.pack(pady=5)
         type_var.set(str(item_data[2]))

    tk.Label(restock_win, text="Add Quantity:").pack()
    qty_entry_dlg = tk.Entry(restock_win)
    qty_entry_dlg.pack(pady=5)
    qty_entry_dlg.focus_set()
    
    def submit_restock():
        qty_val = qty_entry_dlg.get().strip()
        if not qty_val: return
        try: qty = float(qty_val)
        except: messagebox.showerror("Error","Invalid Quantity"); return
        
        new_unit = None
        new_type = None

        if unit_var.get():
             try: new_unit = float(unit_var.get())
             except: messagebox.showerror("Error","Invalid Unit"); return
        
        if is_coffee: new_type = type_var.get()

        wb=openpyxl.load_workbook(filepath); ws=wb["Inventory"]
        found=False
        for r in ws.iter_rows(min_row=2):
            if r[0].value == item_name:
                if is_bar:
                    r[1].value = new_unit
                    r[2].value = (r[2].value or 0) + qty
                    r[4].value = (r[4].value or 0) + qty
                    r[5].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                elif is_coffee:
                    r[1].value = new_unit
                    r[2].value = new_type
                    r[3].value = (r[3].value or 0) + qty
                    r[5].value = (r[5].value or 0) + qty
                    r[6].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                found=True; break
        
        if found:
            wb.save(filepath); wb.close(); load_inventory_table(tree, filepath)
            messagebox.showinfo("Success",f"{item_name} restocked by {qty}")
            restock_win.destroy()
        else: wb.close(); messagebox.showerror("Error", "Item not found")

    tk.Button(restock_win, text="Confirm Restock", bg="#28a745", fg="white", command=submit_restock).pack(pady=15)


def clear_stock_item(tree, filepath):
    sel=tree.selection()
    if not sel: messagebox.showwarning("Error","Select item to clear stock"); return
    item=tree.item(sel[0])['values'][0]
    qty=tk.simpledialog.askinteger("Clear Stock","Enter amount to remove:",minvalue=1)
    if not qty: return
    
    is_bar = "bar" in filepath.lower()
    is_coffee = "coffee" in filepath.lower()
    
    wb=openpyxl.load_workbook(filepath); ws=wb["Inventory"]
    for r in ws.iter_rows(min_row=2):
        if r[0].value==item:
            current = 0
            if is_bar: current=r[2].value
            elif is_coffee: current=r[3].value
            
            if qty > current: messagebox.showerror("Error", f"Cannot remove {qty}, only {current} in stock"); wb.close(); return
            
            if is_bar:
                r[2].value -= qty
                r[5].value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            elif is_coffee:
                r[3].value -= qty
                r[6].value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    wb.save(filepath); wb.close(); load_inventory_table(tree, filepath)
    messagebox.showinfo("Success",f"{item} stock reduced by {qty}")


# ---------------- POS LOGIC ----------------
def add_to_cart():
    item=item_var.get(); qty=qty_entry.get()
    if item=="" or qty=="": messagebox.showwarning("Error","Select item and qty"); return
    try: qty=int(qty); assert qty>0
    except: messagebox.showerror("Error","Quantity must be positive int"); return
    price=MENU[item]; total=price*qty
    cart.append((item,qty,price,total))
    cart_list.insert(tk.END,f"{item} x{qty} = Rs {total}"); update_total()

def update_total(): total_label.config(text=f"Total: Rs {sum(x[3] for x in cart)}")

def generate_bill(auto_print=False):
    global last_receipt_file
    if not cart: messagebox.showwarning("Error","Cart empty"); return
    now=datetime.datetime.now(); bill_id=now.strftime("%Y%m%d%H%M%S")
    total_amount=sum(x[3] for x in cart)
    
    # Save Order
    wb=openpyxl.load_workbook(ORDER_FILE); ws=wb["Orders"]
    for i in cart: ws.append([bill_id, now.strftime("%Y-%m-%d %H:%M:%S"), i[0], i[1], i[2], i[3]])
    wb.save(ORDER_FILE); wb.close()
    
    # Update Inventories
    update_stock_across_inventories(cart)
    
    # Generate Receipt Text
    receipt=[ "        T&T ACADEMY", "  SMART BAR POS SYSTEM", "-------------------------", f"BillID:{bill_id}", f"Date:{now.strftime('%Y-%m-%d %H:%M:%S')}", "-------------------------"]
    for i in cart: receipt.append(f"{i[0]} x{i[1]} = Rs {i[3]}")
    receipt+=["-------------------------", f"TOTAL: Rs {total_amount}", "-------------------------","Thank you!"]
    receipt_text="\n".join(receipt); receipt_box.delete("1.0",tk.END); receipt_box.insert(tk.END,receipt_text)
    
    last_receipt_file=os.path.join(RECEIPT_DIR,f"{bill_id}.txt")
    with open(last_receipt_file,"w") as f: f.write(receipt_text)
    
    cart.clear(); cart_list.delete(0,tk.END); update_total()
    if auto_print: os.system(f'notepad /p "{last_receipt_file}"')
    messagebox.showinfo("Saved","Bill generated and recorded"); load_orders_table()

def update_stock_across_inventories(sold_items):
    for item, qty, _, _ in sold_items:
        deducted = False
        
        # Try Bar Inventory (Count based)
        wb_bar = openpyxl.load_workbook(INVENTORY_BAR_FILE); ws_bar = wb_bar["Inventory"]
        for r in ws_bar.iter_rows(min_row=2):
            if r[0].value == item:
                # r[2] is Stock (Count)
                current_stock = r[2].value or 0
                r[2].value = current_stock - qty 
                r[3].value = (r[3].value or 0) + qty
                r[5].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                deducted = True
                break
        wb_bar.save(INVENTORY_BAR_FILE); wb_bar.close()
        
        if deducted: continue 
        
        # Try Coffee Inventory (Count based now, with Unit Metadata)
        wb_coffee = openpyxl.load_workbook(INVENTORY_COFFEE_FILE); ws_coffee = wb_coffee["Inventory"]
        
        for r in ws_coffee.iter_rows(min_row=2):
            if r[0].value == item:
                # New Schema: Item[0], UnitSize[1], UnitType[2], Stock[3], Used[4]...
                r[3].value = (r[3].value or 0) - qty
                r[4].value = (r[4].value or 0) + qty
                r[6].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                deducted = True
                break
        wb_coffee.save(INVENTORY_COFFEE_FILE); wb_coffee.close()
        
    load_inventory_table(inv_bar_tree, INVENTORY_BAR_FILE)
    load_inventory_table(inv_coffee_tree, INVENTORY_COFFEE_FILE)

def print_bill(): 
    if not last_receipt_file: messagebox.showwarning("Error","Generate bill first"); return
    os.system(f'notepad /p "{last_receipt_file}"')

def clear_all(): cart.clear(); cart_list.delete(0,tk.END); receipt_box.delete("1.0",tk.END); update_total()

# ... (Student and Payment functions unchanged) ...
# ---------------- STUDENTS ----------------
def save_student():
    name=name_entry.get().strip(); contact=contact_entry.get().strip(); loc=location_entry.get().strip()
    total=total_entry.get().strip(); paid=paid_entry.get().strip(); mode=payment_mode_var.get(); remark=remark_entry.get().strip()
    if not name or not contact or not loc or not total or not paid or mode=="Select": messagebox.showwarning("Error","All fields required"); return
    try: total=float(total); paid=float(paid); pending=total-paid; 
    except: messagebox.showerror("Error","Total/Paid must be number"); return
    if pending<0: messagebox.showerror("Error","Paid> Total"); return
    wb=openpyxl.load_workbook(STUDENT_FILE); ws=wb["Students"]; ws.append([name,contact,loc,total,paid,pending,mode,remark]); wb.save(STUDENT_FILE); wb.close()
    load_students_table(); clear_student_form(); messagebox.showinfo("Saved",f"Student {name} saved!")

def clear_student_form(): 
    for e in [name_entry,contact_entry,location_entry,total_entry,paid_entry,remark_entry]: e.delete(0,tk.END)
    pending_entry.config(state="normal"); pending_entry.delete(0,tk.END); pending_entry.config(state="readonly")
    payment_mode_var.set("Select")

def update_pending(*args):
    try: total=float(total_entry.get()); paid=float(paid_entry.get()); pending_entry.config(state="normal"); pending_entry.delete(0,tk.END); pending_entry.insert(0,str(total-paid)); pending_entry.config(state="readonly")
    except: pending_entry.config(state="normal"); pending_entry.delete(0,tk.END); pending_entry.config(state="readonly")

def load_students_table():
    for r in student_table.get_children(): student_table.delete(r)
    wb=openpyxl.load_workbook(STUDENT_FILE); ws=wb["Students"]
    for row in ws.iter_rows(min_row=2,values_only=True): student_table.insert("",tk.END,values=row)
    wb.close()

# ---------------- PAYMENT UPDATE ----------------
def load_student_payment_form():
    global selected_student_row_idx
    sel=student_table.selection(); 
    if not sel: messagebox.showwarning("Error","Select student"); return
    data=student_table.item(sel[0])['values']
    payment_name_var.set(data[0]); payment_total_var.set(data[3]); payment_paid_var.set(data[4]); payment_pending_var.set(data[5])
    payment_mode_var.set(data[6]); remark_entry_payment.delete(0,tk.END); remark_entry_payment.insert(0,data[7]); add_paid_entry.delete(0,tk.END)
    wb=openpyxl.load_workbook(STUDENT_FILE); ws=wb["Students"]
    for idx,row in enumerate(ws.iter_rows(min_row=2),start=2):
        if row[0].value==data[0] and str(row[1].value)==str(data[1]): selected_student_row_idx=idx; break
    wb.close()

def update_payment_amount(*args):
    try: new=float(add_paid_entry.get()); pending=float(payment_total_var.get())-(float(payment_paid_var.get())+new); payment_pending_var.set(pending)
    except: payment_pending_var.set("Error")

def save_updated_payment():
    global selected_student_row_idx
    if not selected_student_row_idx: messagebox.showwarning("Error","Load student first"); return
    try: add=float(add_paid_entry.get())
    except: messagebox.showerror("Error","Enter valid number"); return
    mode=payment_mode_var.get(); remark=remark_entry_payment.get().strip()
    if mode=="Select": messagebox.showwarning("Error","Select Payment Mode"); return
    wb=openpyxl.load_workbook(STUDENT_FILE); ws=wb["Students"]; row=ws[selected_student_row_idx]
    new_paid=row[4].value + add
    if new_paid>row[3].value: messagebox.showerror("Error","Paid > Total"); wb.close(); return
    row[4].value=new_paid; row[5].value=row[3].value-new_paid; row[6].value=mode; row[7].value=remark
    wb.save(STUDENT_FILE); wb.close(); load_students_table(); messagebox.showinfo("Updated","Payment & remarks updated successfully!")
    add_paid_entry.delete(0,tk.END); payment_name_var.set(""); payment_total_var.set(""); payment_paid_var.set(""); payment_pending_var.set(""); payment_mode_var.set("Select"); remark_entry_payment.delete(0,tk.END); selected_student_row_idx=None

# ---------------- ORDERS ----------------
def load_orders_table():
    for r in orders_table.get_children(): orders_table.delete(r)
    wb=openpyxl.load_workbook(ORDER_FILE); ws=wb["Orders"]
    for row in ws.iter_rows(min_row=2,values_only=True): orders_table.insert("",tk.END,values=row)
    wb.close()

# ---------------- SEARCH LOGIC ----------------
def filter_table(tree, data, search_term, columns_to_search=(0,)):
    """Generic table filtering function"""
    for r in tree.get_children(): tree.delete(r)
    search_term = search_term.lower()
    for row in data:
        match = False
        for col_idx in columns_to_search:
            if search_term in str(row[col_idx]).lower():
                match = True
                break
        if match:
            tree.insert("", tk.END, values=row)

def search_students(event=None):
    term = student_search_entry.get()
    wb = openpyxl.load_workbook(STUDENT_FILE); ws = wb["Students"]
    data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    filter_table(student_table, data, term, columns_to_search=(0, 1)) # Search Name and Contact

def search_orders(event=None):
    term = orders_search_entry.get()
    wb = openpyxl.load_workbook(ORDER_FILE); ws = wb["Orders"]
    data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    filter_table(orders_table, data, term, columns_to_search=(0, 2)) # Search BillID and Item

def search_menu(event=None):
    term = menu_search_entry.get()
    data = [(k,v) for k,v in MENU.items()]
    filter_table(menu_table, data, term, columns_to_search=(0,))

def search_inventory_bar(event=None):
    term = bar_search_entry.get()
    wb = openpyxl.load_workbook(INVENTORY_BAR_FILE); ws = wb["Inventory"]
    data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    filter_table(inv_bar_tree, data, term, columns_to_search=(0,))

def search_inventory_coffee(event=None):
    term = coffee_search_entry.get()
    wb = openpyxl.load_workbook(INVENTORY_COFFEE_FILE); ws = wb["Inventory"]
    data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    filter_table(inv_coffee_tree, data, term, columns_to_search=(0,))

def quick_find_student():
    term = pay_search_entry.get().strip().lower()
    if not term: return
    wb = openpyxl.load_workbook(STUDENT_FILE); ws = wb["Students"]
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if term in str(row[0]).lower() or term in str(row[1]).lower():
             # Load this student
             payment_name_var.set(row[0]); payment_total_var.set(row[3]); payment_paid_var.set(row[4]); payment_pending_var.set(row[5])
             payment_mode_var.set(row[6]); remark_entry_payment.delete(0,tk.END); remark_entry_payment.insert(0,row[7] if row[7] else ""); add_paid_entry.delete(0,tk.END)
             global selected_student_row_idx
             selected_student_row_idx = idx
             wb.close()
             messagebox.showinfo("Found", f"Loaded student: {row[0]}")
             return
    wb.close()
    messagebox.showwarning("Not Found", "Student not found")

# ---------------- GUI ----------------
root=tk.Tk(); root.title("T&T ACADEMY - Full POS System"); root.geometry("1400x850"); root.resizable(True,True)
header=tk.Label(root,text="T&T ACADEMY - Smart POS System",bg="#1f1f1f",fg="white",font=("Helvetica",16,"bold"),pady=10); header.pack(fill=tk.X)
notebook=ttk.Notebook(root); notebook.pack(expand=True,fill="both")

# ---------- POS TAB ----------
tab1=tk.Frame(notebook); notebook.add(tab1,text="POS System")
main_frame=tk.Frame(tab1); main_frame.pack(fill="both",expand=True,pady=10)
left_frame=tk.Frame(main_frame,padx=15); left_frame.pack(side=tk.LEFT,fill="y")
right_frame=tk.Frame(main_frame,padx=15); right_frame.pack(side=tk.LEFT,fill="both",expand=True)

tk.Label(left_frame,text="Menu",font=("Helvetica",12)).pack(anchor="w")
item_var=tk.StringVar()
item_menu=tk.OptionMenu(left_frame,item_var,""); item_menu.config(width=25); item_menu.pack()
tk.Label(left_frame,text="Quantity",font=("Helvetica",12)).pack(anchor="w",pady=(10,0))
qty_entry=tk.Entry(left_frame); qty_entry.pack(fill=tk.X)
tk.Button(left_frame,text="Add to Cart",bg="#28a745",fg="white",command=add_to_cart).pack(fill=tk.X,pady=5)
cart_list=tk.Listbox(left_frame,width=40,height=10); cart_list.pack()
total_label=tk.Label(left_frame,text="Total: Rs 0",font=("Helvetica",12,"bold")); total_label.pack(pady=5)
tk.Button(left_frame,text="Generate Bill",bg="#007bff",fg="white",command=lambda: generate_bill(auto_print=True)).pack(fill=tk.X)
tk.Button(left_frame,text="Print Bill",bg="#ffc107",fg="black",command=print_bill).pack(fill=tk.X,pady=3)
tk.Button(left_frame,text="Clear Cart",bg="#dc3545",fg="white",command=clear_all).pack(fill=tk.X)

tk.Label(right_frame,text="Receipt",font=("Helvetica",12,"bold")).pack()
receipt_box=tk.Text(right_frame,width=60,height=15,bg="#f5f5f5"); receipt_box.pack(pady=5,fill="both",expand=True)

search_order_frame=tk.Frame(right_frame); search_order_frame.pack(fill=tk.X, pady=(10,0))
tk.Label(search_order_frame, text="Search Orders:").pack(side=tk.LEFT)
orders_search_entry=tk.Entry(search_order_frame)
orders_search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
orders_search_entry.bind("<KeyRelease>", search_orders)

orders_frame=tk.Frame(right_frame); orders_frame.pack(fill=tk.BOTH,expand=True)
columns_order=("BillID","Date","Item","Qty","Price","Total"); orders_table=ttk.Treeview(orders_frame,columns=columns_order,show="headings",height=7)
for c in columns_order: orders_table.heading(c,text=c); orders_table.column(c,width=100)
scrollbar_orders=ttk.Scrollbar(orders_frame,orient="vertical",command=orders_table.yview); orders_table.configure(yscroll=scrollbar_orders.set); scrollbar_orders.pack(side=tk.RIGHT,fill=tk.Y); orders_table.pack(fill=tk.BOTH,expand=True)

# ---------- STUDENT TAB ----------
tab2=tk.Frame(notebook); notebook.add(tab2,text="Student Management")
student_form=tk.Frame(tab2,pady=10); student_form.pack(side=tk.TOP,fill=tk.X,padx=10)
tk.Label(student_form,text="Name").grid(row=0,column=0,padx=5,pady=5,sticky="w")
name_entry=tk.Entry(student_form,width=25); name_entry.grid(row=0,column=1,padx=5,pady=5)
tk.Label(student_form,text="Contact").grid(row=0,column=2,padx=5,pady=5,sticky="w")
contact_entry=tk.Entry(student_form,width=25); contact_entry.grid(row=0,column=3,padx=5,pady=5)

# Search Student (Top Right corner of same row)
student_form.grid_columnconfigure(4, weight=1)
tk.Label(student_form, text="", font=("Helvetica", 10)).grid(row=0, column=5, padx=5, sticky="e")
student_search_entry=tk.Entry(student_form, width=20)
student_search_entry.grid(row=0, column=6, padx=5, sticky="w")
tk.Button(student_form, text="Search", bg="#007bff", fg="white", command=search_students).grid(row=0, column=7, padx=5, sticky="w")
student_search_entry.bind("<KeyRelease>", search_students)

tk.Label(student_form,text="Location").grid(row=1,column=0,padx=5,pady=5,sticky="w")
location_entry=tk.Entry(student_form,width=25); location_entry.grid(row=1,column=1,padx=5,pady=5)
tk.Label(student_form,text="Total Amount").grid(row=1,column=2,padx=5,pady=5,sticky="w")
total_entry=tk.Entry(student_form,width=25); total_entry.grid(row=1,column=3,padx=5,pady=5)
tk.Label(student_form,text="Paid Amount").grid(row=2,column=0,padx=5,pady=5,sticky="w")
paid_entry=tk.Entry(student_form,width=25); paid_entry.grid(row=2,column=1,padx=5,pady=5)
tk.Label(student_form,text="Pending Amount").grid(row=2,column=2,padx=5,pady=5,sticky="w")
pending_entry=tk.Entry(student_form,width=25,state="readonly"); pending_entry.grid(row=2,column=3,padx=5,pady=5)
payment_mode_var=tk.StringVar(value="Select"); tk.Label(student_form,text="Payment Mode").grid(row=3,column=0,padx=5,pady=5,sticky="w")
tk.OptionMenu(student_form,payment_mode_var,"Cash","Bank","UPI","Card").grid(row=3,column=1,padx=5,pady=5,sticky="w")
tk.Label(student_form,text="Remarks").grid(row=3,column=2,padx=5,pady=5,sticky="w"); remark_entry=tk.Entry(student_form,width=25); remark_entry.grid(row=3,column=3,padx=5,pady=5)
tk.Button(student_form,text="Save Student",bg="#28a745",fg="white",command=save_student).grid(row=4,column=0,columnspan=2,pady=5)
tk.Button(student_form,text="Clear Form",bg="#dc3545",fg="white",command=clear_student_form).grid(row=4,column=2,columnspan=2,pady=5)

student_table_frame=tk.Frame(tab2); student_table_frame.pack(fill=tk.BOTH,expand=True,padx=10)
columns_student=("Name","Contact","Location","Total","Paid","Pending","Payment Mode","Remarks")
student_table=ttk.Treeview(student_table_frame,columns=columns_student,show="headings",height=10)
for c in columns_student: student_table.heading(c,text=c); student_table.column(c,width=100)
scrollbar_student=ttk.Scrollbar(student_table_frame,orient="vertical",command=student_table.yview); student_table.configure(yscroll=scrollbar_student.set); scrollbar_student.pack(side=tk.RIGHT,fill=tk.Y); student_table.pack(fill=tk.BOTH,expand=True)

# ---------- PAYMENT UPDATE TAB ----------
tab3=tk.Frame(notebook); notebook.add(tab3,text="Update Payment")

# Update Payment Search
search_pay_frame = tk.Frame(tab3, padx=10)
search_pay_frame.pack(fill=tk.X, pady=(10,0))
tk.Label(search_pay_frame, text="Quick Find Student (Name/Contact):").pack(side=tk.LEFT)
pay_search_entry = tk.Entry(search_pay_frame)
pay_search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
tk.Button(search_pay_frame, text="Find & Load", bg="#007bff", fg="white", command=quick_find_student).pack(side=tk.LEFT)

payment_frame=tk.Frame(tab3,pady=10); payment_frame.pack(side=tk.TOP,fill=tk.X,padx=10)
payment_name_var=tk.StringVar(); payment_total_var=tk.StringVar(); payment_paid_var=tk.StringVar(); payment_pending_var=tk.StringVar()
tk.Label(payment_frame,text="Name").grid(row=0,column=0,padx=5,pady=5,sticky="w"); tk.Entry(payment_frame,textvariable=payment_name_var,state="readonly").grid(row=0,column=1,padx=5,pady=5)
tk.Label(payment_frame,text="Total").grid(row=0,column=2,padx=5,pady=5,sticky="w"); tk.Entry(payment_frame,textvariable=payment_total_var,state="readonly").grid(row=0,column=3,padx=5,pady=5)
tk.Label(payment_frame,text="Paid").grid(row=1,column=0,padx=5,pady=5,sticky="w"); tk.Entry(payment_frame,textvariable=payment_paid_var,state="readonly").grid(row=1,column=1,padx=5,pady=5)
tk.Label(payment_frame,text="Pending").grid(row=1,column=2,padx=5,pady=5,sticky="w"); tk.Entry(payment_frame,textvariable=payment_pending_var,state="readonly").grid(row=1,column=3,padx=5,pady=5)
tk.Label(payment_frame,text="Add Paid").grid(row=2,column=0,padx=5,pady=5,sticky="w"); add_paid_entry=tk.Entry(payment_frame); add_paid_entry.grid(row=2,column=1,padx=5,pady=5); add_paid_entry.bind("<KeyRelease>",update_payment_amount)
tk.Label(payment_frame,text="Payment Mode").grid(row=2,column=2,padx=5,pady=5,sticky="w"); tk.OptionMenu(payment_frame,payment_mode_var,"Cash","Bank","UPI","Card").grid(row=2,column=3,padx=5,pady=5)
tk.Label(payment_frame,text="Remarks").grid(row=3,column=0,padx=5,pady=5,sticky="w"); remark_entry_payment=tk.Entry(payment_frame,width=50); remark_entry_payment.grid(row=3,column=1,columnspan=3,padx=5,pady=5)
tk.Button(payment_frame,text="Load Selected Student",bg="#007bff",fg="white",command=load_student_payment_form).grid(row=4,column=0,columnspan=2,pady=5)
tk.Button(payment_frame,text="Save Payment Update",bg="#28a745",fg="white",command=save_updated_payment).grid(row=4,column=2,columnspan=2,pady=5)

# ---------- MENU TAB ----------
tab4=tk.Frame(notebook); notebook.add(tab4,text="Menu Management")
menu_form=tk.Frame(tab4,pady=10); menu_form.pack(side=tk.TOP,fill=tk.X,padx=10)
tk.Label(menu_form,text="Item Name").grid(row=0,column=0,padx=5,pady=5); menu_name_entry=tk.Entry(menu_form); menu_name_entry.grid(row=0,column=1,padx=5,pady=5)
tk.Label(menu_form,text="Item Price").grid(row=0,column=2,padx=5,pady=5); menu_price_entry=tk.Entry(menu_form); menu_price_entry.grid(row=0,column=3,padx=5,pady=5)

# Search Menu (Top Right corner of same row)
menu_form.grid_columnconfigure(4, weight=1)
tk.Label(menu_form, text="", font=("Helvetica", 10)).grid(row=0, column=5, padx=5, sticky="e")
menu_search_entry=tk.Entry(menu_form, width=20)
menu_search_entry.grid(row=0, column=6, padx=5, sticky="w")
tk.Button(menu_form, text="Search", bg="#007bff", fg="white", command=search_menu).grid(row=0, column=7, padx=5, sticky="w")
menu_search_entry.bind("<KeyRelease>", search_menu)

tk.Button(menu_form,text="Add Item",bg="#28a745",fg="white",command=add_item_menu).grid(row=1,column=0,padx=5,pady=5)
tk.Button(menu_form,text="Edit Item",bg="#ffc107",fg="black",command=edit_item_menu).grid(row=1,column=1,padx=5,pady=5)
tk.Button(menu_form,text="Remove Item",bg="#dc3545",fg="white",command=remove_item_menu).grid(row=1,column=2,padx=5,pady=5)

menu_table_frame=tk.Frame(tab4); menu_table_frame.pack(fill=tk.BOTH,expand=True,padx=10)
columns_menu=("Item","Price"); menu_table=ttk.Treeview(menu_table_frame,columns=columns_menu,show="headings",height=10)
for c in columns_menu: menu_table.heading(c,text=c); menu_table.column(c,width=150)
scrollbar_menu=ttk.Scrollbar(menu_table_frame,orient="vertical",command=menu_table.yview); menu_table.configure(yscroll=scrollbar_menu.set); scrollbar_menu.pack(side=tk.RIGHT,fill=tk.Y); menu_table.pack(fill=tk.BOTH,expand=True)

# ---------- BAR INVENTORY TAB ----------
tab5=tk.Frame(notebook); notebook.add(tab5,text="Inventory (Bar)")
bar_frame=tk.Frame(tab5); bar_frame.pack(fill="both",expand=True, padx=10, pady=5)

# Bar Search (Top Right)
search_bar_frame=tk.Frame(bar_frame); search_bar_frame.pack(side=tk.TOP, anchor="e", pady=5)
tk.Label(search_bar_frame, text="").pack(side=tk.LEFT)
bar_search_entry=tk.Entry(search_bar_frame); bar_search_entry.pack(side=tk.LEFT, padx=5)
tk.Button(search_bar_frame, text="Search", bg="#007bff", fg="white", command=search_inventory_bar).pack(side=tk.LEFT)
bar_search_entry.bind("<KeyRelease>", search_inventory_bar)

# Bar CRUD Form
bar_crud_frame=tk.Frame(bar_frame); bar_crud_frame.pack(side=tk.TOP,fill=tk.X)
tk.Label(bar_crud_frame,text="Item:").pack(side=tk.LEFT)
bar_name_entry=tk.Entry(bar_crud_frame); bar_name_entry.pack(side=tk.LEFT, padx=5)

# UNIT FIELD
tk.Label(bar_crud_frame,text="Unit (ml):").pack(side=tk.LEFT)
bar_unit_entry=tk.Entry(bar_crud_frame, width=10); bar_unit_entry.pack(side=tk.LEFT, padx=5)

tk.Label(bar_crud_frame,text="Stock:").pack(side=tk.LEFT)
bar_stock_entry=tk.Entry(bar_crud_frame, width=10); bar_stock_entry.pack(side=tk.LEFT, padx=5)
tk.Button(bar_crud_frame,text="Add",bg="#28a745",fg="white",command=lambda: add_inventory_item(inv_bar_tree, INVENTORY_BAR_FILE, bar_name_entry, bar_stock_entry, bar_unit_entry)).pack(side=tk.LEFT, padx=2)
tk.Button(bar_crud_frame,text="Update",bg="#ffc107",fg="black",command=lambda: edit_inventory_item(inv_bar_tree, INVENTORY_BAR_FILE, bar_name_entry, bar_stock_entry, bar_unit_entry)).pack(side=tk.LEFT, padx=2)
tk.Button(bar_crud_frame,text="Delete",bg="#dc3545",fg="white",command=lambda: delete_inventory_item(inv_bar_tree, INVENTORY_BAR_FILE)).pack(side=tk.LEFT, padx=2)


inv_bar_tree=ttk.Treeview(bar_frame,columns=("Item","Unit (ml)","Current Stock","Used","Restocked","Last Updated"),show="headings",height=10)
for c in ("Item","Unit (ml)","Current Stock","Used","Restocked","Last Updated"): inv_bar_tree.heading(c,text=c); inv_bar_tree.column(c,width=120)
bar_scroll=ttk.Scrollbar(bar_frame,orient="vertical",command=inv_bar_tree.yview); inv_bar_tree.configure(yscroll=bar_scroll.set); bar_scroll.pack(side=tk.RIGHT,fill=tk.Y); inv_bar_tree.pack(fill=tk.BOTH,expand=True)
bar_action_frame=tk.Frame(bar_frame); bar_action_frame.pack(pady=5)
tk.Button(bar_action_frame,text="Restock Selected Item",bg="#28a745",fg="white",command=lambda: restock_inventory_item(inv_bar_tree,INVENTORY_BAR_FILE)).pack(side=tk.LEFT, padx=5)
tk.Button(bar_action_frame,text="Stock Clearance Item",bg="#dc3545",fg="white",command=lambda: clear_stock_item(inv_bar_tree,INVENTORY_BAR_FILE)).pack(side=tk.LEFT, padx=5)



# ---------- COFFEE INVENTORY TAB ----------
tab6=tk.Frame(notebook); notebook.add(tab6,text="Inventory (Coffee)")
coffee_frame=tk.Frame(tab6); coffee_frame.pack(fill="both",expand=True, padx=10, pady=5)

# Coffee Search (Top Right)
search_coffee_frame=tk.Frame(coffee_frame); search_coffee_frame.pack(side=tk.TOP, anchor="e", pady=5)
tk.Label(search_coffee_frame, text="").pack(side=tk.LEFT)
coffee_search_entry=tk.Entry(search_coffee_frame); coffee_search_entry.pack(side=tk.LEFT, padx=5)
tk.Button(search_coffee_frame, text="Search", bg="#007bff", fg="white", command=search_inventory_coffee).pack(side=tk.LEFT)
coffee_search_entry.bind("<KeyRelease>", search_inventory_coffee)

# Coffee CRUD Form
coffee_crud_frame=tk.Frame(coffee_frame); coffee_crud_frame.pack(side=tk.TOP,fill=tk.X)
tk.Label(coffee_crud_frame,text="Item:").pack(side=tk.LEFT)
coffee_name_entry=tk.Entry(coffee_crud_frame); coffee_name_entry.pack(side=tk.LEFT, padx=5)

# UNIT FIELDS
tk.Label(coffee_crud_frame,text="Unit:").pack(side=tk.LEFT)
coffee_unit_entry=tk.Entry(coffee_crud_frame, width=8); coffee_unit_entry.pack(side=tk.LEFT, padx=5)
coffee_type_var=tk.StringVar(value="ml")
tk.OptionMenu(coffee_crud_frame, coffee_type_var, "ml", "gm").pack(side=tk.LEFT)

tk.Label(coffee_crud_frame,text="Stock:").pack(side=tk.LEFT)
coffee_stock_entry=tk.Entry(coffee_crud_frame, width=10); coffee_stock_entry.pack(side=tk.LEFT, padx=5)

tk.Button(coffee_crud_frame,text="Add",bg="#28a745",fg="white",command=lambda: add_inventory_item(inv_coffee_tree, INVENTORY_COFFEE_FILE, coffee_name_entry, coffee_stock_entry, coffee_unit_entry, coffee_type_var)).pack(side=tk.LEFT, padx=2)
tk.Button(coffee_crud_frame,text="Update",bg="#ffc107",fg="black",command=lambda: edit_inventory_item(inv_coffee_tree, INVENTORY_COFFEE_FILE, coffee_name_entry, coffee_stock_entry, coffee_unit_entry, coffee_type_var)).pack(side=tk.LEFT, padx=2)
tk.Button(coffee_crud_frame,text="Delete",bg="#dc3545",fg="white",command=lambda: delete_inventory_item(inv_coffee_tree, INVENTORY_COFFEE_FILE)).pack(side=tk.LEFT, padx=2)

inv_coffee_tree=ttk.Treeview(coffee_frame,columns=("Item","Unit","Unit Type(ml/gm)","Current Stock","Used","Restocked","Last Updated"),show="headings",height=10)
for c in ("Item","Unit","Unit Type(ml/gm)","Current Stock","Used","Restocked","Last Updated"): inv_coffee_tree.heading(c,text=c); inv_coffee_tree.column(c,width=120)
coffee_scroll=ttk.Scrollbar(coffee_frame,orient="vertical",command=inv_coffee_tree.yview); inv_coffee_tree.configure(yscroll=coffee_scroll.set); coffee_scroll.pack(side=tk.RIGHT,fill=tk.Y); inv_coffee_tree.pack(fill=tk.BOTH,expand=True)
coffee_action_frame=tk.Frame(coffee_frame); coffee_action_frame.pack(pady=5)
tk.Button(coffee_action_frame,text="Restock Selected Item",bg="#28a745",fg="white",command=lambda: restock_inventory_item(inv_coffee_tree,INVENTORY_COFFEE_FILE)).pack(side=tk.LEFT, padx=5)
tk.Button(coffee_action_frame,text="Stock Clearance Item",bg="#dc3545",fg="white",command=lambda: clear_stock_item(inv_coffee_tree,INVENTORY_COFFEE_FILE)).pack(side=tk.LEFT, padx=5)


# ---------------- INIT ----------------
init_files()
load_menu()
load_students_table()
load_orders_table()
load_inventory_table(inv_bar_tree, INVENTORY_BAR_FILE)
load_inventory_table(inv_coffee_tree, INVENTORY_COFFEE_FILE)

total_entry.bind("<KeyRelease>",update_pending)
paid_entry.bind("<KeyRelease>",update_pending)

root.mainloop()